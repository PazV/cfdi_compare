#--*-- coding: utf-8 --*--
from flask import Flask, render_template, flash, redirect, url_for, session, request, logging, Blueprint, g, send_file
from passlib.hash import sha256_crypt
from functools import wraps
from werkzeug.utils import secure_filename
import logging
# from .login import is_logged_in
import json
import sys
import traceback
import os
import glob
import re
import shutil
import random
import zipfile
from .db_connection import getDB
db = getDB()
from flask import current_app as app
from . import general_functions
GF = general_functions.GeneralFunctions()
import app_config as cfg
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, colors, PatternFill, Border, Alignment, Side, NamedStyle
from openpyxl.cell import Cell
from time import gmtime, strftime, localtime
import datetime
import csv

bp = Blueprint('cfdi', __name__, url_prefix='/cfdi' )

@bp.route('/load-company-file')
# @is_logged_in
def loadCompanyFile():
    return render_template('load_company_file.html',g=g)

@bp.route('/load-sat-file')
# @is_logged_in
def loadSatFile():
    return render_template('load_sat_file.html', g=g)

@bp.route('/compare')
# @is_logged_in
def compareRecords():
    return render_template('compare_records.html', g=g)

@bp.route('/my-progress')
# @is_logged_in
def comparingProgress():
    return render_template('comparing_progress.html', g=g)

@bp.route('/month-detail/<year>/<month>')
# @is_logged_in
def monthDetail(year,month):
    return render_template('progress_month_detail.html',g=g)


@bp.route('/createLoadingFormat', methods=['GET','POST'])
# @is_logged_in
def createLoadingFormat():
    response={}
    try:
        if request.method=='POST':
            headers=['uuid','rfc_emisor','nombre_emisor','rfc_receptor','nombre_receptor','fecha_emision','monto','tipo_comprobante','estatus','fecha_cancelacion']
            wb = Workbook()
            ws = wb.create_sheet("Formato")
            column=1
            for x in headers:
                ws.cell(row=1,column=column,value=x).font=Font(name='Arial',size=12)
                column+=1

            for column_cells in ws.columns:
                length = max(len(GF.as_text(cell.value))+5 for cell in column_cells)
                ws.column_dimensions[column_cells[0].column].width = length


            wb.remove(wb['Sheet'])
            wb.save('%sFormato_carga.xlsx'%cfg.path_for_downloads)
            path='%sFormato_carga.xlsx'%cfg.path_for_downloads
            name='Formato_carga.xlsx'
            response['success']=True
            response['filename']=name
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'

    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/downloadFile/<filename>', methods=['GET','POST'])
# @is_logged_in
def downloadFile(filename):
    response={}
    try:
        path="%s%s"%(cfg.path_for_downloads,filename)
        name="%s"%filename
        return send_file(path,attachment_filename=name)
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
        return json.dumps(response)

@bp.route('/checkCompanyExistingRecords', methods=['GET','POST'])
# @is_logged_in
def checkCompanyExistingRecords():
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                response['success']=True
                exists=db.query("""
                    select *
                    from system.company_cfdi_my_rel
                    where year=%s
                    and company_id=%s
                """%(data['year'],data['company_id'])).dictresult()
                if exists!=[]:
                    months=eval(exists[0]['months'])
                    if months[data['month']]==True:
                        response['exists']=True
                        response['msg_confirm']='Ya existen registros para el mes seleccionado, ¿desea reemplazarlos?'
                    else:
                        response['exists']=False
                else:
                    response['exists']=False
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar validar la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/loadCompanyInfo', methods=['GET','POST'])
# @is_logged_in
def loadCompanyInfo():
    response={}
    try:
        if request.method=='POST':
            # valid,data=GF.getDict(request.form,'post')
            data=request.form.to_dict()
            files=request.files
            file_path=cfg.path_for_downloads
            file=files[data['file_name']]
            filename = secure_filename(file.filename)
            file.save(os.path.join(file_path,filename))
            read_file=load_workbook(os.path.join(file_path,filename))
            ws=read_file.worksheets[0]

            cfdi_list=[]
            row=2

            for r in range(1,ws.max_row):
                record={}
                for x in range(1,int(ws.max_column)+1):
                    if ws.cell(row=row,column=x).value!=None:
                        record[ws.cell(row=1,column=x).value]=ws.cell(row=row,column=x).value
                cfdi_list.append(record)
                row+=1

            #busca si ya existe registro de esa empresa en ese año
            register_exists=db.query("""
                select * from system.company_cfdi_my_rel
                where company_id=%s and year=%s
            """%(data['company_id'],data['year'])).dictresult()
            if data['exists']=='create':
                #crear tabla
                db.query("""
                    CREATE TABLE company_cfdi.c_%s_%s_%s(
                        c_id serial not null primary key,
                        uuid text not null default '',
                        rfc_emisor text not null default '',
                        nombre_emisor text not null default '',
                        rfc_receptor text not null default '',
                        nombre_receptor text not null default '',
                        fecha_emision timestamp without time zone default '1900-01-01 00:00:00',
                        monto float not null default 0,
                        tipo_comprobante text not null default '',
                        estatus text not null default '',
                        fecha_cancelacion timestamp without time zone not null default '1900-01-01 00:00:00'
                    );
                """%(data['company_id'],data['month'],data['year']))

                #inserta los registros a la tabla correspondiente
                for c in cfdi_list:
                    c['estatus']=str(c['estatus']).lower()
                    db.insert('company_cfdi.c_%s_%s_%s'%(data['company_id'],data['month'],data['year']),c)

                if register_exists==[]: #si no existe registro, se crea uno
                    months={
                        'ene':False, 'feb':False, 'mar':False,
                        'abr':False, 'may':False, 'jun':False,
                        'jul':False, 'ago':False, 'sep':False,
                        'oct':False, 'nov':False, 'dic':False
                    }
                    updated={
                        'ene':'1900-01-01 00:00:00', 'feb':'1900-01-01 00:00:00', 'mar':'1900-01-01 00:00:00',
                        'abr':'1900-01-01 00:00:00', 'may':'1900-01-01 00:00:00', 'jun':'1900-01-01 00:00:00',
                        'jul':'1900-01-01 00:00:00', 'ago':'1900-01-01 00:00:00', 'sep':'1900-01-01 00:00:00',
                        'oct':'1900-01-01 00:00:00', 'nov':'1900-01-01 00:00:00', 'dic':'1900-01-01 00:00:00'
                    }
                    months[data['month']]=True
                    updated[data['month']]=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    rel_reg={
                        'company_id':data['company_id'],
                        'year':data['year'],
                        'months':str(months),
                        'updated':str(updated)
                    }
                    #agrega registro de la tabla creada
                    db.insert('system.company_cfdi_my_rel',rel_reg)
                else: #si ya existe registro, se edita para agregar el mes que se está procesando
                    months=eval(register_exists[0]['months'])
                    months[data['month']]=True
                    updated=eval(register_exists[0]['updated'])
                    updated[data['month']]=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    db.query("""
                        update system.company_cfdi_my_rel
                        set months='%s',
                        updated='%s'
                        where company_id=%s and year=%s
                    """%(str(months).replace("'","''"),str(updated).replace("'","''"),data['company_id'],data['year']))

                response['msg_response']='Los registros han sido guardados.'

            elif data['exists']=='replace':
                #obtener el id del último registro de la tabla
                last_id=db.query("""
                    select max(c_id) from company_cfdi.c_%s_%s_%s
                """%(data['company_id'],data['month'],data['year'])).dictresult()[0]['max']

                #inserta los registros a la tabla correspondiente
                for c in cfdi_list:
                    c['estatus']=str(c['estatus']).lower()
                    db.insert('company_cfdi.c_%s_%s_%s'%(data['company_id'],data['month'],data['year']),c)

                #elimina datos de la tabla
                db.query("""
                    delete from company_cfdi.c_%s_%s_%s where c_id < %s
                """%(data['company_id'],data['month'],data['year'],int(last_id)+1))

                updated=eval(register_exists[0]['updated'])
                updated[data['month']]=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                db.query("""
                    update system.company_cfdi_my_rel
                    set updated='%s'
                    where company_id=%s and year=%s
                """%(str(updated).replace("'","''"),data['company_id'],data['year']))

                response['msg_response']='Los registros han sido reemplazados.'
            response['success']=True

        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/checkSatExistingRecords', methods=['GET','POST'])
# @is_logged_in
def checkSatExistingRecords():
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                response['success']=True
                exists=db.query("""
                    select *
                    from system.sat_cfdi_my_rel
                    where year=%s
                    and company_id=%s
                """%(data['year'],data['company_id'])).dictresult()
                if exists!=[]:
                    months=eval(exists[0]['months'])
                    if months[data['month']]==True:
                        response['exists']=True
                        response['msg_confirm']='Ya existen registros para el mes seleccionado, ¿desea reemplazarlos?'
                    else:
                        response['exists']=False
                else:
                    response['exists']=False
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar validar la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)


@bp.route('/loadSatInfo', methods=['GET','POST'])
# @is_logged_in
def loadSatInfo():
    response={}
    try:
        if request.method=='POST':
            #obtiene data
            data=request.form.to_dict()
            #lee archivo
            files=request.files
            file_path=cfg.path_for_downloads
            file=files[data['file_name']]
            filename = secure_filename(file.filename)
            file.save(os.path.join(file_path,filename))

            #convierte el archivo .txt en .csv
            with open (os.path.join(file_path,filename), 'r') as f:
                stripped=(line.strip() for line in f)
                lines=(line.split("~") for line in stripped if line)
                csv_file_name='sat_load.csv'
                with open(os.path.join(file_path,csv_file_name), 'w') as of:
                    writer=csv.writer(of)
                    writer.writerows(lines)

            #revisar si ya hay registro
            existing_record=db.query("""
                select * from system.sat_cfdi_my_rel
                where year=%s and company_id=%s
            """%(data['year'],data['company_id'])).dictresult()
            if data['exists']=='create':
                #crear tabla
                db.query("""
                    CREATE TABLE sat_cfdi.sc_%s_%s_%s(
                        sc_id serial not null primary key,
                        uuid text not null default '',
                        rfc_emisor text not null default '',
                        nombre_emisor text not null default '',
                        rfc_receptor text not null default '',
                        nombre_receptor text not null default '',
                        fecha_emision timestamp without time zone default '1900-01-01 00:00:00',
                        monto float not null default 0,
                        tipo_comprobante text not null default '',
                        estatus text not null default '',
                        fecha_cancelacion timestamp without time zone default '1900-01-01 00:00:00'
                    );
                """%(data['company_id'],data['month'],data['year']))

                ##inserta los registros a la tabla correspondiente
                #lee archivo csv para insertar registros en tabla correspondiente
                with open(os.path.join(file_path,csv_file_name),'rb') as csvfile:
                    fieldnames=['uuid','rfc_emisor','nombre_emisor','rfc_receptor','nombre_receptor','rfc_pac','fecha_emision','fecha_certificacion_sat','monto','tipo_comprobante','estatus','fecha_cancelacion']
                    sreader=csv.DictReader(csvfile,fieldnames=fieldnames, delimiter=',', quotechar='"')
                    count_row=0
                    for row in sreader:

                        if count_row!=0:
                            # app.logger.info(row)
                            row['estatus']=str(row['estatus']).lower()
                            db.insert('sat_cfdi.sc_%s_%s_%s'%(data['company_id'],data['month'],data['year']),row)
                        count_row+=1


                if existing_record==[]: #si no existe registro, se crea uno
                    months={
                        'ene':False, 'feb':False, 'mar':False,
                        'abr':False, 'may':False, 'jun':False,
                        'jul':False, 'ago':False, 'sep':False,
                        'oct':False, 'nov':False, 'dic':False
                    }
                    updated={
                        'ene':'1900-01-01 00:00:00', 'feb':'1900-01-01 00:00:00', 'mar':'1900-01-01 00:00:00',
                        'abr':'1900-01-01 00:00:00', 'may':'1900-01-01 00:00:00', 'jun':'1900-01-01 00:00:00',
                        'jul':'1900-01-01 00:00:00', 'ago':'1900-01-01 00:00:00', 'sep':'1900-01-01 00:00:00',
                        'oct':'1900-01-01 00:00:00', 'nov':'1900-01-01 00:00:00', 'dic':'1900-01-01 00:00:00'
                    }
                    months[data['month']]=True
                    updated[data['month']]=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    rel_reg={
                        'company_id':data['company_id'],
                        'year':data['year'],
                        'months':str(months),
                        'updated':str(updated)
                    }
                    #agrega registro de la tabla creada
                    db.insert('system.sat_cfdi_my_rel',rel_reg)
                else: #si ya existe registro, se edita para agregar el mes que se está procesando
                    months=eval(existing_record[0]['months'])
                    months[data['month']]=True
                    updated=eval(existing_record[0]['updated'])
                    updated[data['month']]=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    db.query("""
                        update system.sat_cfdi_my_rel
                        set months='%s',
                        updated='%s'
                        where company_id=%s and year=%s
                    """%(str(months).replace("'","''"),str(updated).replace("'","''"),data['company_id'],data['year']))
                response['msg_response']='Los registros han sido guardados.'

            elif data['exists']=='replace':
                #obtener el id del último registro de la tabla
                last_id=db.query("""
                    select max(sc_id) from sat_cfdi.sc_%s_%s_%s
                """%(data['company_id'],data['month'],data['year'])).dictresult()[0]['max']

                ##inserta los registros a la tabla correspondiente
                #lee archivo csv para insertar registros en tabla correspondiente
                with open(os.path.join(file_path,csv_file_name),'rb') as csvfile:
                    fieldnames=['uuid','rfc_emisor','nombre_emisor','rfc_receptor','nombre_receptor','rfc_pac','fecha_emision','fecha_certificacion_sat','monto','tipo_comprobante','estatus','fecha_cancelacion']
                    sreader=csv.DictReader(csvfile,fieldnames=fieldnames, delimiter=',', quotechar='"')
                    count_row=0
                    for row in sreader:
                        if count_row!=0:
                            row['estatus']=str(row['estatus']).lower()
                            db.insert('sat_cfdi.sc_%s_%s_%s'%(data['company_id'],data['month'],data['year']),row)
                        count_row+=1

                #elimina datos anteriores de la tabla
                db.query("""
                    delete from sat_cfdi.sc_%s_%s_%s
                    where sc_id < %s
                """%(data['company_id'],data['month'],data['year'],int(last_id)+1))

                updated=eval(existing_record[0]['updated'])
                updated[data['month']]=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                db.query("""
                    update system.sat_cfdi_my_rel
                    set updated='%s'
                    where company_id=%s and year=%s
                """%(str(updated).replace("'","''"),data['company_id'],data['year']))

                response['msg_response']='Los registros han sido actualizados.'

            response['success']=True

        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/checkDataToCompare', methods=['GET','POST'])
# @is_logged_in
def checkDataToCompare():
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                response['available']=False
                response['success']=True
                company=db.query("""
                    select * from system.company_cfdi_my_rel
                    where company_id=%s and year=%s
                """%(data['company_id'],data['year'])).dictresult()
                if company!=[]:
                    company_months=eval(company[0]['months'])
                    if company_months[data['month']]==True:
                        sat=db.query("""
                            select * from system.sat_cfdi_my_rel
                            where company_id=%s and year=%s
                        """%(data['company_id'],data['year'])).dictresult()
                        if sat!=[]:
                            sat_months=eval(sat[0]['months'])
                            if sat_months[data['month']]==True:
                                response['available']=True
                                response['msg_response']='Se puede realizar el comparativo.'
                            else:
                                response['msg_response']='No existen datos del SAT correspondientes al mes seleccionado en el a&ntilde;o %s.'%data['year']
                        else:
                            response['msg_response']='No existe ning&uacute;n registro del SAT correspondiente al a&ntilde;o %s.'%data['year']
                    else:
                        response['msg_response']='No existen datos cargados por la empresa correspondientes al mes seleccionado en el a&ntilde;o %s.'%data['year']
                else:
                    response['msg_response']='No existe ning&uacute;n registro del SAT correspondiente al a&ntilde;o %s.'%data['year']
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar validar los datos capturados.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/doComparison', methods=['GET','POST'])
# @is_logged_in
def doComparison():
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                table_name="c_%s_%s_%s"%(data['company_id'],data['month'],data['year'])
                #consulta registros que se encuentran en ambas tablas
                repeated=db.query("""
                    select a.uuid as a_uuid,
                    a.rfc_receptor,
                    a.nombre_receptor,
                    a.estatus as a_estatus,
                    b.estatus as b_estatus,
                    case
                        when b.estatus='1' and a.estatus='vigente' then 'Igual'
                        when b.estatus='0' and a.estatus='cancelado' then 'Igual'
                        when b.estatus='1' and a.estatus='cancelado' then 'Diferente'
                        when b.estatus='0' and a.estatus='vigente' then 'Diferente'
                        when b.estatus='1' and a.estatus='1' then 'Igual'
                    	when b.estatus='0' and a.estatus='0' then 'Igual'
                    	when b.estatus='1' and a.estatus='0' then 'Diferente'
                    	when b.estatus='0' and a.estatus='1' then 'Diferente'
                    else '--' end as comparacion_estatus,
                    case when a.estatus='0' then a.fecha_cancelacion else '1900-01-01' end as fecha_cancelacion,
                    a.monto as a_monto,
                    b.monto as b_monto,
                    (a.monto-b.monto) as diferencia_monto,
                    to_char(a.fecha_emision,'DD-MM-YYYY HH:MI:SS') as a_fe,
                    to_char(b.fecha_emision,'DD-MM-YYYY HH:MI:SS') as b_fe,
                    case when to_char(a.fecha_emision,'DD-MM-YYYY')=to_char(b.fecha_emision,'DD-MM-YYYY') then 'Igual' else 'Diferente' end as fe_dif
                    from company_cfdi.%s a
                    inner join sat_cfdi.s%s b on lower(a.uuid)=lower(b.uuid)
                """%(table_name,table_name)).dictresult()

                #consulta registros que se encuentran solo en la tabla del sat
                only_sat=db.query("""
                    select a.uuid as a_uuid,
                    a.rfc_receptor,
                    a.nombre_receptor,
                    a.estatus as a_estatus,
                    a.monto as a_monto,
                    to_char(a.fecha_emision,'DD-MM-YYYY HH:MI:SS') as a_fe
                    from sat_cfdi.s%s a
                    left join company_cfdi.%s b on lower(a.uuid)=lower(b.uuid)  where b.uuid is NULL;
                """%(table_name,table_name)).dictresult()

                #Crear libro de excel
                wb = Workbook()
                ws = wb.create_sheet('Empresa-SAT',0)

                #Crear estilos para hojas de excel
                header_style=NamedStyle(name='header_style')
                header_style.font=Font(
                    name='Arial',
                    size=12,
                    bold=True,
                    italic=False,
                    color='FFFFFFFF'
                )
                header_style.fill=PatternFill(
                    "solid",
                    fgColor="ff0066b3"
                )
                header_style.alignment=Alignment(
                    horizontal='center'
                )

                content_style=NamedStyle(name='content_style')
                content_style.font=Font(
                    name='Arial',
                    size=10,
                    bold=False,
                    italic=False,
                    color='FF000000'
                )
                content_style.alignment=Alignment(
                    horizontal='justify',
                    vertical='center',
                    text_rotation=0,
                    wrap_text=True,
                    shrink_to_fit=False,
                    indent=0
                )

                st_left_style=NamedStyle(name='st_left_style')
                st_left_style.font=Font(
                    # name='Arial',
                    # size=12,
                    # bold=True,
                    # italic=False,
                    # color='FF000000'
                    name='Arial',
                    size=12,
                    bold=True,
                    italic=False,
                    color='FFFFFFFF'
                )
                st_left_style.alignment=Alignment(
                    horizontal='right',
                    vertical='center',
                    text_rotation=0,
                    wrap_text=False,
                    shrink_to_fit=False,
                    indent=0
                )
                st_left_style.fill=PatternFill(
                    "solid",
                    fgColor="ff0066b3"
                )
                st_right_style=NamedStyle(name='st_right_style')
                st_right_style.font=Font(
                    name='Arial',
                    size=12,
                    bold=False,
                    italic=True,
                    color='FF000000'
                )
                st_right_style.alignment=Alignment(
                    horizontal='center',
                    vertical='center',
                    text_rotation=0,
                    wrap_text=False,
                    shrink_to_fit=False,
                    indent=0
                )

                #Crear hoja principal del comparativo
                header_titles=['UUID','RFC Receptor','Nombre Receptor','Estatus SAT','Estatus Empresa','Comparación Estatus','Fecha Cancelación SAT','Total SAT','Total Empresa','Diferencia Total','Fecha Emisión SAT','Fecha Emisión Empresa', 'Comparación Fecha Emisión']
                ht_count=1
                for ht in header_titles:
                    ws.cell(column=ht_count,row=1,value=ht)
                    ws.cell(column=ht_count,row=1).style=header_style
                    ht_count+=1

                r_count=2
                count_estatus_dif=0
                fe_dif=0
                monto_dif=0
                for r in repeated:
                    ws.cell(column=1,row=r_count,value=r['a_uuid'])
                    ws.cell(column=2,row=r_count,value=r['rfc_receptor'])
                    ws.cell(column=3,row=r_count,value=r['nombre_receptor'])
                    ws.cell(column=4,row=r_count,value=r['b_estatus'])
                    ws.cell(column=5,row=r_count,value=r['a_estatus'])
                    ws.cell(column=6,row=r_count,value=r['comparacion_estatus'])
                    ws.cell(column=7,row=r_count,value=r['fecha_cancelacion'])
                    ws.cell(column=8,row=r_count,value=r['b_monto'])
                    ws.cell(column=9,row=r_count,value=r['a_monto'])
                    ws.cell(column=10,row=r_count,value=r['diferencia_monto'])
                    ws.cell(column=11,row=r_count,value=r['b_fe'])
                    ws.cell(column=12,row=r_count,value=r['a_fe'])
                    ws.cell(column=13,row=r_count,value=r['fe_dif'])
                    if r['comparacion_estatus']=='Diferente':
                        count_estatus_dif+=1
                    if r['fe_dif']=='Diferente':
                        fe_dif+=1
                    if float(r['diferencia_monto'])>0:
                        monto_dif+=1

                    for a in range(1,14):
                        ws.cell(column=a,row=r_count).style=content_style


                    r_count+=1

                #Crear hoja con datos que solo se encontraron en el SAT
                headers_os=['UUID','RFC Receptor','Nombre Receptor','Estatus','Total','Fecha Emisión']
                ws2 = wb.create_sheet('Solo SAT',1)
                ht_count2=1
                for ht2 in headers_os:
                    ws2.cell(column=ht_count2,row=1,value=ht2)
                    ws2.cell(column=ht_count2,row=1).style=header_style
                    ht_count2+=1

                os_count=2
                for on_s in only_sat:
                    ws2.cell(column=1,row=os_count,value=on_s['a_uuid'])
                    ws2.cell(column=2,row=os_count,value=on_s['rfc_receptor'])
                    ws2.cell(column=3,row=os_count,value=on_s['nombre_receptor'])
                    ws2.cell(column=4,row=os_count,value=on_s['a_estatus'])
                    ws2.cell(column=5,row=os_count,value=on_s['a_monto'])
                    ws2.cell(column=6,row=os_count,value=on_s['a_fe'])
                    os_count+=1

                total_sat=db.query("""
                    select count(*) from sat_cfdi.s%s
                """%table_name).dictresult()[0]['count']
                total_company=db.query("""
                    select count(*) from company_cfdi.%s
                """%table_name).dictresult()[0]['count']
                total_repeated=len(repeated)

                #Crear hoja con estadísticas del reporte
                meses={
                    'ene':'Enero','feb':'Febrero','mar':'Marzo','abr':'Abril','may':'Mayo','jun':'Junio','jul':'Julio',
                    'ago':'Agosto','sep':'Septiembre','oct':'Octubre','nov':'Noviembre','dic':'Diciembre'
                }
                ws3=wb.create_sheet('Estad.',2)
                values_left=['Mes analizado:','Total CFDI registrados en portal del SAT:','Total CFDI registrados por la empresa:','Total CFDI encontrados en ambos registros:','Total registros encontrados con estatus diferente:','Total registros encontrados con fecha de emisión diferentes:','Total registros encontrados con diferencia en totales:']
                count_row_st=2
                for vl in values_left:
                    ws3.cell(column=2,row=count_row_st,value=vl)
                    ws3.cell(column=2,row=count_row_st).style=st_left_style
                    count_row_st+=1

                comp_info=[]
                count_row_st=2
                values_right=["%s - %s"%(meses[data['month']],data['year']),total_sat,total_company,total_repeated,count_estatus_dif,fe_dif,monto_dif]
                for vr in values_right:
                    ws3.cell(column=3,row=count_row_st,value=vr)
                    ws3.cell(column=3,row=count_row_st).style=st_right_style
                    comp_info.append('%s %s'%(values_left[count_row_st-2],vr))
                    count_row_st+=1


                #Ajustar ancho de columnas
                for w in wb.sheetnames:
                    sheet=wb[w]
                    for column_cells in sheet.columns:
                        length = max(len(GF.as_text(cell.value))+5 for cell in column_cells)
                        sheet.column_dimensions[column_cells[0].column].width = length

                #Asignar nombre al archivo y guardarlo
                time=strftime("%H_%M_%S",gmtime())
                file_name='Rel-%s-%s_%s.xlsx'%(data['month'],data['year'],time)
                path=os.path.join(cfg.path_for_downloads,file_name)
                wb.save(path)

                response['msg_response']='El comparativo ha sido realizado con éxito, ¿desea descargar el reporte asociado?'
                response['filename']='/cfdi/downloadFile/%s'%file_name

                exists_cd=db.query("""
                    select * from system.comparison_data
                    where year=%s and month='%s' and company_id=%s
                """%(data['year'],data['month'],data['company_id'])).dictresult()
                if exists_cd==[]:
                    comparison_data={
                        'company_id':data['company_id'],
                        'month':data['month'],
                        'year':data['year'],
                        'comparison_date':'now',
                        'info':str(comp_info)
                    }
                    db.insert('system.comparison_data',comparison_data)
                else:
                    db.query("""
                        update system.comparison_data
                        set info='%s',
                        comparison_date='now'
                        where company_id=%s and month='%s' and year=%s
                    """%(str(comp_info).replace("'","''"),data['company_id'],data['month'],data['year']))
                response['success']=True
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar validar la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/getMonthsProgress', methods=['GET','POST'])
# @is_logged_in
def getMonthsProgress():
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                company=db.query("""
                    select months
                    from system.company_cfdi_my_rel
                    where company_id=%s and year=%s
                """%(data['company_id'],data['year'])).dictresult()

                sat=db.query("""
                    select months
                    from system.sat_cfdi_my_rel
                    where company_id=%s and year=%s
                """%(data['company_id'],data['year'])).dictresult()

                do_for=False
                if company!=[] and sat!=[]:
                    c_months=eval(company[0]['months'])
                    s_months=eval(sat[0]['months'])
                    do_for=True
                else:
                    if company!=[] or sat!=[]:
                        do_for=True
                        if company==[]:
                            c_months={'ene':False,'feb':False,'mar':False,'abr':False,'may':False,'jun':False,'jul':False,'ago':False,'sep':False,'oct':False,'nov':False,'dic':False}
                        else:
                            c_months=eval(company[0]['months'])
                        if sat==[]:
                            s_months={'ene':False,'feb':False,'mar':False,'abr':False,'may':False,'jun':False,'jul':False,'ago':False,'sep':False,'oct':False,'nov':False,'dic':False}
                        else:
                            s_months=eval(company[0]['months'])
                    else:
                        do_for=False
                months={}
                if do_for==True:
                    for k,v in c_months.iteritems():
                        comparison=db.query("""
                            select info from system.comparison_data
                            where company_id=%s and year=%s and month='%s'
                        """%(data['company_id'],data['year'],k)).dictresult()
                        if comparison!=[]:
                            months[k]='a-mp-green'
                        else:
                            if v==True and s_months[k]==True:
                                months[k]='a-mp-blue'
                            elif v==True or s_months[k]==True:
                                months[k]='a-mp-orange'
                            else:
                                months[k]='a-mp-gray'
                else:
                    months={'ene':'a-mp-gray','feb':'a-mp-gray','mar':'a-mp-gray','abr':'a-mp-gray','may':'a-mp-gray','jun':'a-mp-gray','jul':'a-mp-gray','ago':'a-mp-gray','sep':'a-mp-gray','oct':'a-mp-gray','nov':'a-mp-gray','dic':'a-mp-gray'}
                response['success']=True
                response['data']=months

            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar validar la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)

@bp.route('/getMonthDetailInfo', methods=['GET','POST'])
# @is_logged_in
def getMonthDetailInfo():
    response={}
    try:
        if request.method=='POST':
            valid,data=GF.getDict(request.form,'post')
            if valid:
                response['success']=True
                exists_comparison=db.query("""
                    select * from system.comparison_data
                    where company_id=%s
                    and year=%s and month='%s'
                """%(data['company_id'],data['year'],data['month'])).dictresult()
                if exists_comparison!=[]:
                    response['comparison']=True
                    response['data']=eval(exists_comparison[0]['info'])
                    response['color_class']='month-bg-green'
                else:
                    response['comparison']=False
                    sat=db.query("""
                        select * from system.sat_cfdi_my_rel
                        where year=%s and company_id=%s
                    """%(data['year'],data['company_id'])).dictresult()
                    company=db.query("""
                        select * from system.company_cfdi_my_rel
                        where company_id=%s and year=%s
                    """%(data['company_id'],data['year'])).dictresult()
                    resp_info={}
                    table_name="c_%s_%s_%s"%(data['company_id'],data['month'],data['year'])
                    if sat!=[]:
                        sat_months=eval(sat[0]['months'])
                        if sat_months[data['month']]==True:
                            sat_date=eval(sat[0]['updated'])
                            sat_total=db.query("""
                                select count(*) from sat_cfdi.s%s
                            """%table_name).dictresult()
                            resp_info['sat']=['CFDIs registrados en el SAT','Estatus: Archivo cargado','Fecha de carga: %s'%sat_date[data['month']],'Total registros cargados: %s'%sat_total[0]['count']]
                        else:
                            resp_info['sat']=['CFDIs registrados en el SAT','Estatus: Pendiente por cargar','Fecha de carga: --','Total registros cargados: --']
                    else:
                        resp_info['sat']=['CFDIs registrados en el SAT','Estatus: Pendiente por cargar','Fecha de carga: --','Total registros cargados: --']
                    if company!=[]:
                        comp_months=eval(company[0]['months'])
                        if comp_months[data['month']]==True:
                            comp_date=eval(company[0]['updated'])
                            comp_total=db.query("""
                                select count(*) from company_cfdi.%s
                            """%table_name).dictresult()
                            resp_info['company']=['CFDIs registrados por la empresa','Estatus: Archivo cargado','Fecha de carga: %s'%comp_date[data['month']],'Total registros cargados: %s'%comp_total[0]['count']]
                        else:
                            resp_info['company']=['CFDIs registrados por la empresa','Esatus: Pendiente por cargar', 'Fecha de carga: --','Total registros cargados: --']
                    else:
                        resp_info['company']=['CFDIs registrados por la empresa','Esatus: Pendiente por cargar', 'Fecha de carga: --','Total registros cargados: --']
                    if company!=[] and sat!=[]:
                        response['color_class']='month-bg-blue'
                    else:
                        if company!=[] or sat!=[]:
                            response['color_class']='month-bg-orange'
                        else:
                            response['color_class']='month-bg-gray'
                    response['data']=resp_info
            else:
                response['success']=False
                response['msg_response']='Ocurrió un error al intentar validar la información.'
        else:
            response['success']=False
            response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo.'
    except:
        response['success']=False
        response['msg_response']='Ocurrió un error, favor de intentarlo de nuevo más tarde.'
        app.logger.info(traceback.format_exc(sys.exc_info()))
    return json.dumps(response)
